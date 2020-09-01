# -*- coding: utf-8 -*-
import Levenshtein as lst
from openpyxl import load_workbook
def read_account_std_excel(std_file):
    wb = load_workbook(std_file)
    l = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            _id = ws.cell(row=r, column=2).value
            _scope = ws.cell(row=r, column=4).value
            dic = dict(
                id=_id,
                name=ws.cell(row=r, column=3).value,
                scope=_scope if _scope else "普遍适用",
                kind=sheet,
            )
            if _id:
                l.append(dic)
    return l
def best_match(target,names):
    max = 0
    best=None
    for name in names:
        ratio = lst.ratio(target, name)
        if ratio > max:
            max = ratio
            best = name
    return best,max
