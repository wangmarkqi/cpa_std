from cpa_std.vehicle_insurance_converter.common.tools import Tool
from PIL import Image
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import shutil,os
from openpyxl.utils import get_column_letter
from pyzbar import pyzbar
class Qr:
    def __init__(self):
        self.t=Tool()
    def get_qr(self,imgs):
        qr_std = "http://eservice"
        for f in imgs:
            img = Image.open(f)
            data = pyzbar.decode(img)
            if len(data) == 0:continue
            data2 = data[0]
            if len(data2)==0:continue
            data3=str(data2[0])
            if qr_std in data3:
                return f
        return  None

    # def index2name(self, row, column):
    #     return get_column_letter(column) + str(row)
    # def get_img(self,ws,f):
    #     l=[]
    #     name = os.path.basename(f).replace(".xlsx","")
    #     for r in range(1, ws.max_row + 1):
    #         for c in range(1, ws.max_column + 1):
    #             image_loader = SheetImageLoader(ws)
    #             id=self.index2name(r,c)
    #             if image_loader.image_in(id):
    #                 image = image_loader.get(id)
    #                 path = f"{self.t.temp}/{name}_{id}.png"
    #                 l.append({
    #                     "path":path,
    #                     "img":image,
    #                 })
    #     if len(l)>0:
    #         self.save_img(l)
    #         return [i['path'] for i in l]
    #     return []
    # def save_img(self,l):
    #     for dic in l:
    #         dic['img'].save(dic['path'])
    #
